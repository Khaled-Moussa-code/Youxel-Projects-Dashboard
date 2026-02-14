"""
Sprint Analysis Automation - Data Processor Module

This module handles Azure DevOps data extraction and processing.
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Tuple
from datetime import datetime


class SprintDataProcessor:
    """Process Azure DevOps sprint data from Excel"""
    
    def __init__(self, workbook_path: str):
        self.workbook_path = workbook_path
        self.team_mapping = {
            'Development': 'Development',
            'Front End': 'Frontend',
            'Design': 'Design',
            'Testing': 'Testing',
            'Business': 'Business'
        }
    
    def extract_sprint_metadata(self, data_sheet) -> Dict:
        """
        Extract sprint metadata from Data sheet rows 3-18
        
        Returns:
            Dict with sprint information
        """
        metadata = {}
        
        # Extract from fixed positions (Excel is 1-indexed, pandas is 0-indexed)
        metadata['sprint_name'] = data_sheet.cell(row=3, column=3).value
        metadata['date_from'] = data_sheet.cell(row=4, column=3).value
        metadata['date_to'] = data_sheet.cell(row=5, column=3).value
        metadata['product'] = data_sheet.cell(row=6, column=3).value
        metadata['escaped_bugs'] = data_sheet.cell(row=7, column=3).value
        metadata['midsprint_addition'] = data_sheet.cell(row=8, column=3).value
        metadata['adhoc'] = data_sheet.cell(row=9, column=3).value
        metadata['planned_sp'] = data_sheet.cell(row=10, column=3).value
        metadata['actual_sp'] = data_sheet.cell(row=11, column=3).value
        metadata['estimated_dev_effort'] = data_sheet.cell(row=12, column=3).value
        metadata['actual_dev_effort'] = data_sheet.cell(row=13, column=3).value
        metadata['estimated_test_effort'] = data_sheet.cell(row=14, column=3).value
        metadata['actual_test_effort'] = data_sheet.cell(row=15, column=3).value
        metadata['bug_fixing_effort'] = data_sheet.cell(row=16, column=3).value
        metadata['sum_actual_hours'] = data_sheet.cell(row=17, column=3).value
        metadata['sum_working_hours'] = data_sheet.cell(row=18, column=3).value
        
        return metadata
    
    def process_azure_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Process Azure DevOps data from DataFrame
        
        Args:
            df: DataFrame with Azure DevOps data (starting from row 21)
        
        Returns:
            Processed DataFrame with additional calculated columns
        """
        # Clean column names
        df.columns = df.columns.str.strip()
        
        # Extract email from 'Assigned To' field
        df['staff_email'] = df['Assigned To'].str.extract(r'<(.+?)>')[0]
        df['staff_name'] = df['Assigned To'].str.extract(r'^(.+?)\s*<')[0]
        
        # Map Activity to Team
        df['team'] = df['Activity'].map(self.team_mapping)
        
        # Handle tags
        df['has_midsprint_addition'] = df['Tags'].fillna('').str.contains('MidSprint_Addition', case=False)
        df['has_adhoc'] = df['Tags'].fillna('').str.contains('Ad-hoc', case=False)
        
        # Convert numeric columns
        numeric_cols = ['Original Estimate', 'Completed Work', 'Remaining Work', 'Story Points']
        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        
        # Calculate task completion
        df['is_done'] = df['State'].isin(['Done', 'Closed'])
        
        return df
    
    def get_capacity_data(self, capacity_sheet) -> pd.DataFrame:
        """
        Extract capacity data from Capacity sheet
        
        Returns:
            DataFrame with capacity by staff member
        """
        capacity_data = []
        
        # Read from row 4 onwards (row 3 has headers)
        for row_num in range(4, 15):  # Adjust based on actual team size
            member = capacity_sheet.cell(row=row_num, column=2).value
            if member is None:
                continue
                
            team = capacity_sheet.cell(row=row_num, column=3).value
            expected_hours = capacity_sheet.cell(row=row_num, column=4).value
            vacation = capacity_sheet.cell(row=row_num, column=5).value or 0
            days_off = capacity_sheet.cell(row=row_num, column=6).value or 0
            remaining_capacity = capacity_sheet.cell(row=row_num, column=7).value
            
            capacity_data.append({
                'staff_member': member,
                'team': team,
                'expected_hours': expected_hours,
                'vacation_hours': vacation,
                'days_off_hours': days_off,
                'remaining_capacity': remaining_capacity
            })
        
        return pd.DataFrame(capacity_data)
    
    def aggregate_by_staff(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Aggregate metrics by staff member
        
        Returns:
            DataFrame with aggregated staff metrics
        """
        staff_agg = df.groupby(['staff_email', 'staff_name', 'team']).agg({
            'Original Estimate': 'sum',
            'Completed Work': 'sum',
            'Remaining Work': 'sum',
            'is_done': ['sum', 'count'],
            'has_midsprint_addition': 'sum',
            'has_adhoc': 'sum'
        }).reset_index()
        
        # Flatten column names
        staff_agg.columns = [
            'staff_email', 'staff_name', 'team',
            'original_estimate_total', 'completed_work_total', 
            'remaining_work_total', 'done_count', 'total_count',
            'midsprint_count', 'adhoc_count'
        ]
        
        return staff_agg
    
    def aggregate_by_team(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Aggregate metrics by team
        
        Returns:
            DataFrame with aggregated team metrics
        """
        team_agg = df.groupby('team').agg({
            'Original Estimate': 'sum',
            'Completed Work': 'sum',
            'Remaining Work': 'sum',
            'is_done': ['sum', 'count'],
            'has_midsprint_addition': 'sum',
            'has_adhoc': 'sum'
        }).reset_index()
        
        # Flatten column names
        team_agg.columns = [
            'team', 'original_estimate_total', 
            'completed_work_total', 'remaining_work_total',
            'done_count', 'total_count',
            'midsprint_count', 'adhoc_count'
        ]
        
        return team_agg
    
    def validate_data(self, df: pd.DataFrame) -> Dict:
        """
        Validate data quality and completeness
        
        Returns:
            Dict with validation results
        """
        validation_results = {
            'status': 'success',
            'errors': [],
            'warnings': []
        }
        
        # Check required columns
        required_cols = [
            'Work Item Type', 'Title', 'State', 'Assigned To',
            'Original Estimate', 'Completed Work', 'Activity'
        ]
        
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            validation_results['status'] = 'error'
            validation_results['errors'].append({
                'type': 'missing_columns',
                'message': f"Missing required columns: {missing_cols}"
            })
        
        # Check for null values in critical fields
        if df['Assigned To'].isna().any():
            validation_results['warnings'].append({
                'type': 'missing_assignments',
                'count': df['Assigned To'].isna().sum(),
                'message': f"{df['Assigned To'].isna().sum()} tasks have no assignee"
            })
        
        # Check for unmapped activities
        unmapped = df[df['team'].isna()]
        if len(unmapped) > 0:
            validation_results['warnings'].append({
                'type': 'unmapped_activities',
                'activities': unmapped['Activity'].unique().tolist(),
                'message': f"Unmapped activities found: {unmapped['Activity'].unique().tolist()}"
            })
        
        # Check for tasks with completed work but no state change
        incomplete_with_work = df[(df['Completed Work'] > 0) & (~df['is_done'])]
        if len(incomplete_with_work) > 0:
            validation_results['warnings'].append({
                'type': 'incomplete_tasks_with_work',
                'count': len(incomplete_with_work),
                'message': f"{len(incomplete_with_work)} tasks have completed work but are not marked as Done"
            })
        
        return validation_results


def main():
    """Example usage"""
    import openpyxl
    
    # Load workbook
    file_path = '/mnt/user-data/uploads/Sprint_25_JAN_to_05_FEB__copy.xlsx'
    wb = openpyxl.load_workbook(file_path, data_only=False)
    
    # Initialize processor
    processor = SprintDataProcessor(file_path)
    
    # Extract sprint metadata
    data_sheet = wb['Data']
    sprint_info = processor.extract_sprint_metadata(data_sheet)
    print("Sprint Info:", sprint_info)
    
    # Process Azure DevOps data
    df_azure = pd.read_excel(file_path, sheet_name='Data', header=20)
    df_processed = processor.process_azure_data(df_azure)
    
    # Validate data
    validation = processor.validate_data(df_processed)
    print("\nValidation Results:", validation)
    
    # Aggregate by staff
    staff_metrics = processor.aggregate_by_staff(df_processed)
    print("\nStaff Metrics Preview:")
    print(staff_metrics.head())
    
    # Aggregate by team
    team_metrics = processor.aggregate_by_team(df_processed)
    print("\nTeam Metrics Preview:")
    print(team_metrics)


if __name__ == "__main__":
    main()
