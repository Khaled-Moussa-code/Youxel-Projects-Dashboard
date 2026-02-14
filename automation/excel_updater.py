"""
Sprint Analysis Automation - Excel Updater Module

This module handles updating Excel sheets with calculated metrics.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd
from typing import Dict, List


class ExcelUpdater:
    """Update Excel sheets with calculated metrics using formulas"""
    
    def __init__(self, workbook_path: str):
        self.workbook_path = workbook_path
        self.wb = openpyxl.load_workbook(workbook_path)
    
    def update_analysis_sheet(self,
                             sprint_name: str,
                             staff_metrics: pd.DataFrame,
                             team_metrics: pd.DataFrame):
        """
        Update or create Analysis sheet with current sprint data
        
        Creates formulas that reference the raw data in Data sheet
        """
        sheet_name = f"{sprint_name} Analysis"
        
        # Create new sheet or clear existing
        if sheet_name in self.wb.sheetnames:
            del self.wb[sheet_name]
        
        sheet = self.wb.create_sheet(sheet_name)
        
        # Title
        sheet['B2'] = f'Sprint Name: {sprint_name}'
        sheet['B2'].font = Font(bold=True, size=14)
        
        # STAFF SECTION
        start_row = 5
        
        # Headers
        headers = ['Row Labels', 'Sum of Original Estimate', 'Sum of Completed Work', 
                   'Sum of Remaining Work', 'Performance Rate', 'Capacity', 'Utilization Per Staff']
        for col_idx, header in enumerate(headers, start=2):
            cell = sheet.cell(row=start_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', fill_type='solid')
        
        # Staff data with formulas
        current_row = start_row + 1
        for idx, staff in staff_metrics.iterrows():
            sheet.cell(row=current_row, column=2, value=staff['staff_name'])
            
            # These reference the Data sheet pivot tables
            # In production, you'd use SUMIFS to pull from Data sheet
            sheet.cell(row=current_row, column=3, value=staff['original_estimate'])
            sheet.cell(row=current_row, column=4, value=staff['completed_work'])
            sheet.cell(row=current_row, column=5, value=staff['remaining_work'])
            
            # Performance Rate formula
            perf_formula = f"=C{current_row}/(D{current_row}+E{current_row})"
            sheet.cell(row=current_row, column=6, value=perf_formula)
            
            # Capacity (could be a VLOOKUP to Capacity sheet)
            sheet.cell(row=current_row, column=7, value=staff['capacity'])
            
            # Utilization formula
            util_formula = f"=D{current_row}/G{current_row}"
            sheet.cell(row=current_row, column=8, value=util_formula)
            
            current_row += 1
        
        # Grand Total row
        total_row = current_row
        sheet.cell(row=total_row, column=2, value='Grand Total')
        sheet.cell(row=total_row, column=2).font = Font(bold=True)
        
        # Sum formulas for totals
        data_start = start_row + 1
        data_end = current_row - 1
        
        sheet.cell(row=total_row, column=3, value=f"=SUM(C{data_start}:C{data_end})")
        sheet.cell(row=total_row, column=4, value=f"=SUM(D{data_start}:D{data_end})")
        sheet.cell(row=total_row, column=5, value=f"=SUM(E{data_start}:E{data_end})")
        sheet.cell(row=total_row, column=6, value=f"=C{total_row}/(D{total_row}+E{total_row})")
        sheet.cell(row=total_row, column=7, value=f"=SUM(G{data_start}:G{data_end})")
        sheet.cell(row=total_row, column=8, value=f"=D{total_row}/G{total_row}")
        
        # TEAM SECTION
        team_start_row = total_row + 3
        
        # Team headers
        for col_idx, header in enumerate(headers, start=2):
            cell = sheet.cell(row=team_start_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', fill_type='solid')
        
        # Team data
        current_row = team_start_row + 1
        for idx, team in team_metrics.iterrows():
            sheet.cell(row=current_row, column=2, value=team['team'])
            sheet.cell(row=current_row, column=3, value=team['original_estimate'])
            sheet.cell(row=current_row, column=4, value=team['completed_work'])
            sheet.cell(row=current_row, column=5, value=team['remaining_work'])
            
            # Formulas
            sheet.cell(row=current_row, column=6, value=f"=C{current_row}/(D{current_row}+E{current_row})")
            sheet.cell(row=current_row, column=7, value=team['capacity'])
            sheet.cell(row=current_row, column=8, value=f"=D{current_row}/G{current_row}")
            
            current_row += 1
        
        # Format columns
        for col in ['C', 'D', 'E', 'G']:
            sheet.column_dimensions[col].width = 20
        
        sheet.column_dimensions['B'].width = 40
        sheet.column_dimensions['F'].width = 18
        sheet.column_dimensions['H'].width = 22
    
    def update_kpi_indicators_sheet(self,
                                    staff_metrics: pd.DataFrame,
                                    team_metrics: pd.DataFrame,
                                    sprint_name: str):
        """
        Update KPI Indicators sheet with formulas referencing Analysis sheet
        """
        if 'Kpi Indicators' not in self.wb.sheetnames:
            sheet = self.wb.create_sheet('Kpi Indicators')
        else:
            sheet = self.wb['Kpi Indicators']
        
        analysis_sheet_name = f"{sprint_name} Analysis"
        
        # STAFF KPI SECTION
        sheet['B2'] = 'Row Labels'
        sheet['C2'] = 'Performance Rate'
        sheet['D2'] = 'Weight'
        sheet['E2'] = 'Utilization Per Staff'
        sheet['F2'] = 'Weight'
        sheet['G2'] = 'Done Tasks %'
        sheet['H2'] = 'Weight'
        sheet['I2'] = 'KPI'
        
        # Format headers
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            sheet[f'{col}2'].font = Font(bold=True)
        
        # Staff KPIs
        row = 3
        for idx, staff in staff_metrics.iterrows():
            sheet.cell(row=row, column=2, value=staff['staff_name'])
            
            # Reference to Analysis sheet (row numbers would need to match)
            analysis_row = row + 48  # Offset to match Analysis sheet structure
            
            # Performance Rate from Analysis
            sheet.cell(row=row, column=3, value=f"='{analysis_sheet_name}'!F{analysis_row}")
            sheet.cell(row=row, column=4, value=0.2)  # Weight
            
            # Utilization from Analysis
            sheet.cell(row=row, column=5, value=f"='{analysis_sheet_name}'!H{analysis_row}")
            sheet.cell(row=row, column=6, value=0.1)  # Weight
            
            # Done Tasks % (would need to calculate from Data sheet)
            sheet.cell(row=row, column=7, value=staff['done_tasks_pct'])
            sheet.cell(row=row, column=8, value=0.7)  # Weight
            
            # KPI formula
            sheet.cell(row=row, column=9, value=f"=(C{row}*D{row})+(E{row}*F{row})+(G{row}*H{row})")
            
            row += 1
        
        # Grand Total
        sheet.cell(row=row, column=2, value='Grand Total')
        sheet.cell(row=row, column=2).font = Font(bold=True)
        
        data_start = 3
        data_end = row - 1
        
        sheet.cell(row=row, column=3, value=f"=AVERAGE(C{data_start}:C{data_end})")
        sheet.cell(row=row, column=4, value=0.2)
        sheet.cell(row=row, column=5, value=f"=AVERAGE(E{data_start}:E{data_end})")
        sheet.cell(row=row, column=6, value=0.1)
        sheet.cell(row=row, column=7, value=f"=AVERAGE(G{data_start}:G{data_end})")
        sheet.cell(row=row, column=8, value=0.7)
        sheet.cell(row=row, column=9, value=f"=(C{row}*D{row})+(E{row}*F{row})+(G{row}*H{row})")
        
        # TEAM KPI SECTION
        team_start = row + 3
        
        sheet.cell(row=team_start, column=2, value='Team')
        for col, header in zip(['C', 'D', 'E', 'F', 'G', 'H', 'I'],
                              ['Performance Rate', 'Weight', 'Utilization Per Staff', 
                               'Weight', 'Done Tasks %', 'Weight', 'KPI']):
            sheet[f'{col}{team_start}'].value = header
            sheet[f'{col}{team_start}'].font = Font(bold=True)
        
        # Team KPIs
        row = team_start + 1
        for idx, team in team_metrics.iterrows():
            sheet.cell(row=row, column=2, value=team['team'])
            
            # Similar formulas as staff but for teams
            sheet.cell(row=row, column=3, value=team['performance_rate'])
            sheet.cell(row=row, column=4, value=0.2)
            sheet.cell(row=row, column=5, value=team['utilization'])
            sheet.cell(row=row, column=6, value=0.1)
            sheet.cell(row=row, column=7, value=team['done_tasks_pct'])
            sheet.cell(row=row, column=8, value=0.7)
            sheet.cell(row=row, column=9, value=f"=(C{row}*D{row})+(E{row}*F{row})+(G{row}*H{row})")
            
            row += 1
    
    def append_to_historical_staff(self,
                                   staff_metrics: pd.DataFrame,
                                   sprint_name: str):
        """
        Append current sprint data to KPI Indicators Per Staff sheet
        """
        sheet_name = 'Kpi Indicators Per Staff'
        
        if sheet_name not in self.wb.sheetnames:
            sheet = self.wb.create_sheet(sheet_name)
            # Initialize headers
            sheet['B2'] = 'Kpi Indicators Per Staff'
            sheet['B2'].font = Font(bold=True, size=14)
            sheet['B3'] = 'Team Member'
            sheet['B3'].font = Font(bold=True)
        else:
            sheet = self.wb[sheet_name]
        
        # Find next available column
        max_col = sheet.max_column
        next_col = max_col + 1
        
        # Add sprint name as header
        sheet.cell(row=3, column=next_col, value=sprint_name)
        sheet.cell(row=3, column=next_col).font = Font(bold=True)
        
        # Add KPI values
        row = 4
        for idx, staff in staff_metrics.iterrows():
            # Check if staff member exists
            staff_row = None
            for r in range(4, sheet.max_row + 1):
                if sheet.cell(row=r, column=2).value == staff['staff_name']:
                    staff_row = r
                    break
            
            if staff_row is None:
                # Add new staff member
                staff_row = sheet.max_row + 1
                sheet.cell(row=staff_row, column=2, value=staff['staff_name'])
            
            # Add KPI value
            sheet.cell(row=staff_row, column=next_col, value=staff['kpi'])
    
    def append_to_historical_team(self,
                                  team_metrics: pd.DataFrame,
                                  sprint_name: str):
        """
        Append current sprint data to KPI Indicators Per Team sheet
        """
        sheet_name = 'Kpi Indicators Per Team'
        
        if sheet_name not in self.wb.sheetnames:
            sheet = self.wb.create_sheet(sheet_name)
            # Initialize headers
            sheet['B2'] = 'Item'
            sheet['B2'].font = Font(bold=True)
        else:
            sheet = self.wb[sheet_name]
        
        # Find next available column
        max_col = sheet.max_column
        next_col = max_col + 1
        
        # Add sprint name as header
        sheet.cell(row=2, column=next_col, value=sprint_name)
        sheet.cell(row=2, column=next_col).font = Font(bold=True)
        
        # Add team KPI values
        for idx, team in team_metrics.iterrows():
            # Find or create team row
            team_row = None
            for r in range(3, sheet.max_row + 1):
                if sheet.cell(row=r, column=2).value == team['team']:
                    team_row = r
                    break
            
            if team_row is None:
                team_row = sheet.max_row + 1
                sheet.cell(row=team_row, column=2, value=team['team'])
            
            # Add KPI value
            sheet.cell(row=team_row, column=next_col, value=team['kpi'])
    
    def update_cmmi_template(self, cmmi_measures: Dict, sprint_name: str):
        """
        Update CMMI Template sheet with current sprint measures
        """
        sheet_name = 'CMMI Template'
        
        if sheet_name not in self.wb.sheetnames:
            sheet = self.wb.create_sheet(sheet_name)
            # Initialize headers
            sheet['B2'] = 'CMMI Measures History'
            sheet['B2'].font = Font(bold=True, size=14)
            
            headers = ['Sprint', 'Completion Rate', 'Effort Estimation Accuracy',
                      'Bug Fixing %', 'Utilization Rate', 'CMMI Productivity']
            for col_idx, header in enumerate(headers, start=2):
                sheet.cell(row=3, column=col_idx, value=header)
                sheet.cell(row=3, column=col_idx).font = Font(bold=True)
        else:
            sheet = self.wb[sheet_name]
        
        # Find next row
        next_row = sheet.max_row + 1
        
        # Add CMMI data
        sheet.cell(row=next_row, column=2, value=sprint_name)
        sheet.cell(row=next_row, column=3, value=cmmi_measures['completion_rate'])
        sheet.cell(row=next_row, column=4, value=cmmi_measures['effort_estimation_accuracy'])
        sheet.cell(row=next_row, column=5, value=cmmi_measures['bug_fixing_effort_pct'])
        sheet.cell(row=next_row, column=6, value=cmmi_measures['utilization_rate'])
        sheet.cell(row=next_row, column=7, value=cmmi_measures['cmmi_productivity'])
        
        # Format percentages
        for col in [3, 4, 5, 6]:
            sheet.cell(row=next_row, column=col).number_format = '0.00%'
    
    def save(self):
        """Save the workbook"""
        self.wb.save(self.workbook_path)
        print(f"Workbook saved to {self.workbook_path}")


def main():
    """Example usage"""
    import pandas as pd
    
    # Example data
    staff_metrics = pd.DataFrame([
        {
            'staff_name': 'Ahmed Talaat',
            'original_estimate': 89,
            'completed_work': 64,
            'remaining_work': 1,
            'capacity': 63,
            'performance_rate': 0.98,
            'utilization': 1.02,
            'done_tasks_pct': 1.00,
            'kpi': 1.02
        }
    ])
    
    team_metrics = pd.DataFrame([
        {
            'team': 'Development',
            'original_estimate': 200,
            'completed_work': 180,
            'remaining_work': 10,
            'capacity': 189,
            'performance_rate': 0.95,
            'utilization': 0.95,
            'done_tasks_pct': 0.92,
            'kpi': 0.93
        }
    ])
    
    cmmi_measures = {
        'completion_rate': 1.79,
        'effort_estimation_accuracy': 0.85,
        'bug_fixing_effort_pct': 0.0026,
        'utilization_rate': 1.29,
        'cmmi_productivity': 1.39
    }
    
    # Initialize updater
    updater = ExcelUpdater('/mnt/user-data/uploads/Sprint_25_JAN_to_05_FEB__copy.xlsx')
    
    # Update sheets
    updater.update_analysis_sheet('Check Now 26.17', staff_metrics, team_metrics)
    updater.update_kpi_indicators_sheet(staff_metrics, team_metrics, 'Check Now 26.17')
    updater.append_to_historical_staff(staff_metrics, 'Check Now 26.17')
    updater.append_to_historical_team(team_metrics, 'Check Now 26.17')
    updater.update_cmmi_template(cmmi_measures, 'Check Now 26.17')
    
    # Save
    updater.save()


if __name__ == "__main__":
    main()
